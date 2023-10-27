import sys
import pandas as pd
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles import Border, Side
import openpyxl
class GR_main(QMainWindow):
    closed = pyqtSignal()
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Graph_Rename")
        self.setAcceptDrops(True)
        self.resize(500, 200)
        self.path_lb = QLabel("File_path")
        self.path_input = QLineEdit()


        self.save_button = QPushButton("Save and Export")
        self.save_button.clicked.connect(self.save_and_export)
        
        hlayout1=QHBoxLayout()
        hlayout2=QHBoxLayout()
        
        self.graph_lb = QLabel("Title_name : ")
        self.graph_input = QLineEdit()
        self.graph_input .setPlaceholderText("file name1-! file name2")
        hlayout1.addWidget(self.graph_lb)
        hlayout1.addWidget(self.graph_input)
        
        self.y_title = QLabel("Y_title_name : ")
        self.y_title_input = QLineEdit()
        hlayout2.addWidget(self.y_title)
        hlayout2.addWidget(self.y_title_input)
        

        layout = QVBoxLayout()
        layout.addWidget(self.path_lb)
        layout.addWidget(self.path_input)

        layout.addLayout(hlayout1)
        layout.addLayout(hlayout2)
       
        layout.addWidget(self.save_button)
       
        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        for f in files:
            self.path_input.setText(f)

  

    

    def save_and_export(self):
        dialog = Select_window()
        result = dialog.exec_()
        if result == QDialog.Accepted:
            
            file_name=self.chooseFolder()+"/"+dialog.file_name_input.text()+".xlsx"
         
            if dialog.radio_button1.isChecked():
                
                workbook = openpyxl.load_workbook(self.path_input.text())
                wa=workbook.active
                # 시트 내 모든 그래프 이름 변경
                if "!" in self.graph_input.text():
                    new_name= self.graph_input.text().split("!")  # 그래프 이름의 앞에 붙일 접두사
                
                num=1
                for chart in wa._charts:
                    try :
                        chart.title = new_name[0]+str(num)+new_name[1]
                        num+=1
                    except:
                        if self.graph_input.text() !="":
                            chart.title=self.graph_input.text()
                    if self.y_title_input.text() !="":
                        chart.y_axis.title = self.y_title_input.text()
                   
            else:
                workbook = openpyxl.load_workbook(self.path_input.text())
                sheet_names = workbook.sheetnames
                for sheet_name in sheet_names:
                    workbook.active = workbook[sheet_name]
                    wa =workbook.active
                    
                    if "!" in self.graph_input.text():
                        new_name= self.graph_input.text().split("!")
                    num=1
                    for chart in wa._charts:
                        
                        try :
                            chart.title = new_name[0]+str(num)+new_name[1]
                            num+=1
                        except:
                            if self.graph_input.text() !="":
                                chart.title=self.graph_input.text()
                        if self.y_title_input.text() !="":
                            chart.y_axis.title = self.y_title_input.text()
            try :
                workbook.save(file_name)
                workbook.close()
                QMessageBox.about(self,"successfully saved",f'The file has been successfully saved to {self.folder_path}')
            except Exception as e:
                QMessageBox.about(self,f'Error Message',str(e))
                
    def chooseFolder(self):
         options = QFileDialog.Options()
         self.folder_path = QFileDialog.getExistingDirectory(self, "Choose Folder", options=options)
         
         if self.folder_path:
             return self.folder_path
    def closeEvent(self, event):
        self.closed.emit()
class Select_window(QDialog):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.resize(400, 200)
        self.setWindowTitle('Save')
        
        hlayout=QHBoxLayout()
        self.file_name=QLabel("File_Name : ")
        self.file_name_input=QLineEdit()
        
        hlayout.addWidget(self.file_name)
        hlayout.addWidget(self.file_name_input)
        
        hlayout_sheet=QHBoxLayout()
        self.radio_button1 = QRadioButton('one sheet', self)
        self.radio_button1.setChecked(True)
        hlayout_sheet.addWidget(self.radio_button1)

        self.radio_button2 = QRadioButton('All sheet', self)
        hlayout_sheet.addWidget(self.radio_button2)
        
        ok_button = QPushButton('OK', self)
        ok_button.clicked.connect(self.accept)

        cancel_button = QPushButton('Cancel', self)
        cancel_button.clicked.connect(self.reject)
        
        hlayout3 = QHBoxLayout()
        hlayout3.addWidget(ok_button)
        hlayout3.addWidget(cancel_button)
        
        layout=QVBoxLayout()
        layout.addLayout(hlayout_sheet)
        layout.addLayout(hlayout)
        layout.addLayout(hlayout3)
        self.setLayout(layout)
     