import sys
import pandas as pd

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles import Border, Side
from Show_ex import Show_ex

class exm_main(QMainWindow):
    closed = pyqtSignal()  # closed Signal을 정의
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel_Split")
        self.setAcceptDrops(True)
        self.resize(800, 700)
        
        self.path_lb = QLabel("File_path ")
        self.path_input = QLineEdit()
        
        self.sheet=""
        self.dprow = ""
        self.dp = ""
        self.MC=None

        self.table_widget = QTableWidget()
        hl1=QHBoxLayout()
        self.add_row_button = QPushButton("Add Row")
        self.add_row_button.clicked.connect(self.add_row)

        self.add_column_button = QPushButton("Add Column")
        self.add_column_button.clicked.connect(self.add_column)
        hl1.addWidget(self.add_row_button)
        hl1.addWidget(self.add_column_button)
        
        
        hl2=QHBoxLayout()
        self.delete_row_button = QPushButton("Delete Row")
        self.delete_row_button.clicked.connect(self.delete_row)

        self.delete_column_button = QPushButton("Delete Column")
        self.delete_column_button.clicked.connect(self.delete_column)
        hl2.addWidget(self.delete_row_button)
        hl2.addWidget(self.delete_column_button)
        
        self.save_button = QPushButton("Save and Export")
        self.save_button.clicked.connect(self.save_and_export)

        self.ROW_lb = QLabel("Save ROW")
        self.Save_row_input = QLineEdit()

        self.ROW_emp = QLabel("spaces apart")
        self.emp_input = QLineEdit()
        

        layout = QVBoxLayout()
        layout.addWidget(self.path_lb)
        layout.addWidget(self.path_input)

        layout.addWidget(self.table_widget)
        layout.addLayout(hl1)
        layout.addLayout(hl2)
        
        layout.addWidget(self.ROW_lb)
        layout.addWidget(self.Save_row_input)
        layout.addWidget(self.ROW_emp)
        layout.addWidget(self.emp_input)
        layout.addWidget(self.ROW_emp)
        layout.addWidget(self.save_button)
        
        self.mergele=True
        self.mg =[]

        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)
        
    def openDialog(self):
         dialog = Show_ex(self.path_input.text())
         result = dialog.exec_()

         if result == QDialog.Accepted:
             self.sheet = dialog.sheet
             self.dprow = dialog.col
             self.MC=dialog.radio_button1.isChecked()
             if self.MC:
                 self.dp = dialog.value
             else :
                dp, ok = QInputDialog.getText(self, "Excel file name", "name:",text=dialog.value)
                if ok:
                    self.dp=dp
             print(self.sheet ,self.dprow ,  self.dp )
             self.open_new_window()
         else:
             print('Dialog canceled.')
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        for f in files:
            self.path_input.setText(f)
        self.openDialog()
    def open_new_window(self):
        st = 'part'
        nt = 0
        pa = []
        pas = ['part0']
       
        df = pd.read_excel(self.path_input.text(), sheet_name=self.sheet )
       
        if self.MC:
            for i in df[df.columns[int(self.dprow)]]:
                if self.dp == str(i):
                    nt += 1
                    pas.append(st + str(nt))
                pa.append(st + str(nt))
        else:
            for i in df[df.columns[int(self.dprow)]]:
                if self.dp in str(i):
                    nt += 1
                    pas.append(st + str(nt))
                pa.append(st + str(nt)) 
        df['part'] = pa

        dfs = []
        for name in pas:
            fil = df['part'] == name
            i = df.loc[fil].reset_index()
            i = i.drop(['part', 'index'], axis=1)
            dfs.append(i)

        self.dataframes = dfs[1:]
        self.update_table_widget()

    def update_table_widget(self):
        dataframe = self.dataframes[0]
        self.table_widget.setRowCount(dataframe.shape[0])
        self.table_widget.setColumnCount(dataframe.shape[1])
        
        cols=[chr(65+i) for i in range(len(dataframe.columns))]
        
        self.table_widget.setHorizontalHeaderLabels(cols)

        for i in range(dataframe.shape[0]):
            for j in range(dataframe.shape[1]):
                cell_value = str(dataframe.iloc[i, j])
                self.table_widget.setItem(i, j, QTableWidgetItem(cell_value))

    def add_row(self):
        em = ["" for _ in range(len(self.dataframes[0].columns))]
        for i, dataframe in enumerate(self.dataframes):
            df = pd.DataFrame(index=[""], columns=dataframe.columns)
            self.dataframes[i] = pd.concat([df, self.dataframes[i]]).reset_index(drop=True)

        self.update_table_widget()

    def add_column(self):
        column_name, ok = QInputDialog.getText(self, "Add Column", "Enter column name:")
        for i, dataframe in enumerate(self.dataframes):
            if ok and column_name:
                new_column = pd.Series([""] * dataframe.shape[0], dtype=str)
                self.dataframes[i][column_name] = new_column
        self.update_table_widget()

    def delete_row(self):
        selected_rows = sorted(set(index.row() for index in self.table_widget.selectedIndexes()), reverse=True)

        for i in range(len(self.dataframes)):
            self.dataframes[i].drop(self.dataframes[i].index[selected_rows], inplace=True)
            self.dataframes[i] = self.dataframes[i].reset_index(drop=True)
        self.update_table_widget()

    def delete_column(self):
        selected_columns = [index.column() for index in self.table_widget.selectedIndexes()]
        print(selected_columns)
        print(self.dataframes[0].columns[selected_columns])
        for i, dataframe in enumerate(self.dataframes):
            columns_to_delete = dataframe.columns[selected_columns]
            self.dataframes[i].drop(columns_to_delete, axis=1, inplace=True)

        self.update_table_widget()

    def save_and_export(self):
        sr = self.Save_row_input.text().split(",")
        empty = self.emp_input.text()
        modified_dataframe = self.dataframes[0].copy()
        for i in range(self.table_widget.rowCount()):
            for j in range(self.table_widget.columnCount()):
                cell_value = self.table_widget.item(i, j).text()
                if cell_value == "nan":
                    cell_value = ""
                try:
                    cell_value = float(cell_value)
                except:
                    pass
                modified_dataframe.iloc[i, j] = cell_value
        self.dataframes[0] = modified_dataframe
        if sr[0] != '':
            for df in self.dataframes:
                for i in sr:
                    df.loc[int(i)] = self.dataframes[0].loc[int(i)].values

        if empty != "":
            for i in range(len(self.dataframes)):
                for _ in range(int(empty)):
                    self.dataframes[i] = pd.concat([self.dataframes[i], pd.Series()], ignore_index=True)
        merged_dataframe = pd.concat(self.dataframes, axis=0, ignore_index=True)
        merged_dataframe.to_excel("output.xlsx", index=False,header=False)
        
        wb = load_workbook("output.xlsx")
        ws=wb.active
        
        if len(self.mg) != 0:
            border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for val,sr,sc,er,ec in self.mg:
                num=[]
                e=er-sr
                for n in range(1,ws.max_row):
                    if ws[f'{chr(64+sc)}{n}'].value == val:
                          num.append(n)
                for i in num:
                    ws.merge_cells(start_row=i, start_column=sc, end_row=i+e, end_column=ec)
                    for row in ws.iter_rows(min_row=i, min_col=sc, max_row=i+e, max_col=ec):
                        for cell in row:
                            cell.border = border_style
                    ws[f'{chr(64+sc)}{i}'].alignment=Alignment(horizontal='center',vertical='center')
            
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
              for col_index, cell_value in enumerate(row, start=1):
                  if cell_value is not None:
                      cell = ws.cell(row=row_index, column=col_index)
                  cell.border = border_style
            
            try :
                wb.save("result.xlsx")
                QMessageBox.about(self,f'The file has been successfully saved .')
            except Exception as e:
                QMessageBox.about(self,f'Error Message:"{e}".')
                
        
    def merge_st(self):
        row = self.table_widget.currentIndex().row()
        col = self.table_widget.currentIndex().column()
        item = self.table_widget.item(row, col)
        value = item.text()
        
        self.st=[value,row+1,col+1]
        
        self.mergele=False
        
    def merge_en(self):
        row = self.table_widget.currentIndex().row()
        col = self.table_widget.currentIndex().column()
        item = self.table_widget.item(row, col)
        value = item.text()
        
        self.st.append(row+1)
        self.st.append(col+1)
        
        self.mg.append(self.st)
        self.mergele=True
        
    def contextMenuEvent(self, event):
            context_menu = QMenu(self)
            if self.mergele:  
                merge_action = context_menu.addAction("여기부터")
                action = context_menu.exec_(self.mapToGlobal(event.pos()))
                if action != None:
                    self.merge_st()
                    
                
            else:           
                merge_action = context_menu.addAction("여기까지")
                action = context_menu.exec_(self.mapToGlobal(event.pos()))
                if action != None:
                    self.merge_en()

    def closeEvent(self, event):
        self.closed.emit()  # 창이 닫힐 때 'closed' Signal을 발생


